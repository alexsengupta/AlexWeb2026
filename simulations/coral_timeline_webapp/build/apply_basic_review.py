#!/usr/bin/env python3
"""Read the returned review workbook back into the events master.

Run this after editing the blue columns of coral_timeline_basic_text_REVIEW.xlsx.
It only touches the plain-language columns and the further-reading links; the
advanced text, the sources and the projection fields are left exactly as they are.

    python3 build/apply_basic_review.py \
        --master data/events_master.csv \
        --workbook coral_timeline_basic_text_REVIEW.xlsx \
        --out data/events_master.csv

Then re-solve the axis and rebuild:
    python3 build/build_timeline_axis.py --events data/events_master.csv --outdir .
    python3 build/build_corridor.py ...
"""
from __future__ import annotations
import argparse, csv, json
from pathlib import Path
from openpyxl import load_workbook


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", default="data/events_master.csv")
    ap.add_argument("--workbook", default="coral_timeline_basic_text_REVIEW.xlsx")
    ap.add_argument("--out", default="data/events_master.csv")
    args = ap.parse_args()

    ws = load_workbook(args.workbook, data_only=True)["Text"]
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    need = ["Event", "PLAIN headline", "PLAIN what happened", "PLAIN why it matters",
            "Link 1 title", "Link 1 URL", "Link 1 publisher",
            "Link 2 title", "Link 2 URL", "Link 2 publisher"]
    absent = [h for h in need if h not in hdr]
    if absent:
        raise SystemExit("workbook is missing columns: " + ", ".join(absent))

    def cell(r, name):
        v = ws.cell(row=r, column=hdr[name]).value
        return str(v).strip() if v is not None else ""

    edits, warn = {}, []
    for r in range(2, ws.max_row + 1):
        eid = cell(r, "Event")
        if not eid:
            continue
        links = []
        for n in ("1", "2"):
            t, u = cell(r, f"Link {n} title"), cell(r, f"Link {n} URL")
            if u and not u.lower().startswith("http"):
                warn.append(f"{eid} link {n}: not a URL, skipped ({u[:40]})")
                continue
            # a URL with no title, or a title with no URL, is half an edit; say so
            # rather than shipping a blank link or a link with no label
            if bool(t) != bool(u):
                warn.append(f"{eid} link {n}: title and URL must both be filled in, skipped")
                continue
            if t and u:
                links.append({"t": t, "u": u, "p": cell(r, f"Link {n} publisher"), "s": ""})
        edits[eid] = {
            "basic_headline": cell(r, "PLAIN headline"),
            "basic_significance": cell(r, "PLAIN what happened"),
            "basic_why": cell(r, "PLAIN why it matters"),
            "basic_links": json.dumps(links, ensure_ascii=False) if links else "",
        }

    rows = list(csv.DictReader(open(args.master, newline="", encoding="utf-8-sig")))
    fields = list(rows[0].keys())
    changed = blanked = 0
    for row in rows:
        e = edits.get(row["event_id"])
        if not e:
            continue
        for k, v in e.items():
            if k != "basic_links" and not v:
                # never let a blank cell wipe text that is already there
                blanked += 1
                continue
            if row.get(k, "") != v:
                row[k] = v
                changed += 1

    out = Path(args.out)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader(); w.writerows(rows)

    print(f"Wrote {out}")
    print(f"  {len(edits)} rows read from the workbook, {changed} fields changed")
    if blanked:
        print(f"  {blanked} blank cells ignored so existing text was not wiped")
    for w_ in warn:
        print(f"  WARNING: {w_}")


if __name__ == "__main__":
    main()
