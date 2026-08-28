#!/usr/bin/env python3
"""
Apply the manually researched source corrections back into the events master.

Reads the returned follow-up workbook, takes every filled REPLACEMENT DOI / URL,
and rewrites the matching source_N_url in the master. Author corrections that were
recorded in the NOTES column are applied from AUTHOR_FIXES below, which is explicit
rather than parsed out of prose, so a reworded note can never silently change a
citation.

A source whose replacement was left blank could not be sourced at all. Those are
cleared rather than left pointing at a DOI now known to be wrong.

    python3 scripts/apply_source_fixes.py \
        --master coral_reef_events_master_categorised.csv \
        --workbook coral_timeline_source_followup_WITH_URLS.xlsx \
        --out coral_reef_events_master_sourced.csv
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

SHEETS = ["1 Wrong paper", "2 Dead DOI", "3 Tombstone", "4 Author or year off",
          "5 No abstract", "6 Non-DOI sources"]

# Attribution corrections, keyed by the DOI they apply to. Taken from the NOTES
# column but written out here so the change is reviewable and cannot drift.
AUTHOR_FIXES = {
    "10.1038/s41586-025-09615-6": ("Quattrini et al. (2025)", "Vaga et al. (2025)"),
    "10.1016/j.quascirev.2024.108624": ("McNiven et al. (2024)", "Ulm et al. (2024)"),
    "10.1038/s41396-022-01194-y": ("Lawrence et al. (2022)", "Grupstra et al. (2022)"),
    "10.1038/s41467-023-43287-y": ("Quattrini et al. (2023)", "Campoy et al. (2023)"),
}


def as_url(value: str) -> str:
    v = value.strip()
    if not v:
        return ""
    if v.lower().startswith("http"):
        return v
    return "https://doi.org/" + v.lstrip("/")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", default="coral_reef_events_master_categorised.csv")
    ap.add_argument("--workbook", default="coral_timeline_source_followup_WITH_URLS.xlsx")
    ap.add_argument("--out", default="coral_reef_events_master_sourced.csv")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    replace: dict[str, str] = {}   # old DOI -> new URL
    clear: set[str] = set()        # old DOI -> could not be sourced

    for name in SHEETS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        col_doi = hdr.get("DOI")
        col_rep = hdr.get("REPLACEMENT DOI / URL")
        col_src = hdr.get("SOURCE URL")
        if not col_doi:
            continue
        for r in range(2, ws.max_row + 1):
            old = (ws.cell(row=r, column=col_doi).value or "").strip()
            if not old:
                continue
            rep = (ws.cell(row=r, column=col_rep).value or "").strip() if col_rep else ""
            src = (ws.cell(row=r, column=col_src).value or "").strip() if col_src else ""
            new = as_url(rep) or ""
            # SOURCE URL alone confirms the existing DOI; it is only a replacement
            # when it points somewhere different
            if not new and src and src.rstrip("/") != ("https://doi.org/" + old).rstrip("/"):
                new = src
            if new:
                replace[old] = new
            elif name in ("1 Wrong paper", "2 Dead DOI", "3 Tombstone"):
                clear.add(old)

    with open(args.master, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
        fields = list(rows[0].keys())

    n_rep = n_clear = n_auth = 0
    cleared_events = []
    for row in rows:
        for n in ("1", "2"):
            uk, tk = f"source_{n}_url", f"source_{n}_title"
            url = row[uk].strip()
            if "doi.org" not in url:
                continue
            doi = url.split("doi.org/")[-1].strip()
            if doi in replace:
                row[uk] = replace[doi]
                n_rep += 1
                new_doi = replace[doi].split("doi.org/")[-1] if "doi.org" in replace[doi] else ""
                fix = AUTHOR_FIXES.get(new_doi) or AUTHOR_FIXES.get(doi)
                if fix and fix[0] in row[tk]:
                    row[tk] = row[tk].replace(fix[0], fix[1]); n_auth += 1
            elif doi in clear:
                row[uk] = ""
                row[tk] = row[tk] + "  [no citable record found]"
                cleared_events.append(f"{row['event_id']} s{n}")
                n_clear += 1
            else:
                fix = AUTHOR_FIXES.get(doi)
                if fix and fix[0] in row[tk]:
                    row[tk] = row[tk].replace(fix[0], fix[1]); n_auth += 1

    out = Path(args.out)
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader(); w.writerows(rows)

    stranded = [r["event_id"] for r in rows
                if not r["source_1_url"].strip() and not r["source_2_url"].strip()]
    print(f"Wrote {out}")
    print(f"  {len(replace)} distinct DOIs replaced, applied to {n_rep} source slots")
    print(f"  {n_auth} author attributions corrected")
    print(f"  {n_clear} sources cleared as uncitable: {', '.join(cleared_events) or 'none'}")
    print(f"  events left with no source at all: {stranded or 'none'}")


if __name__ == "__main__":
    main()
