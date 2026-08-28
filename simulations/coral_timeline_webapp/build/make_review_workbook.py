#!/usr/bin/env python3
"""Build the side-by-side review workbook for the plain-language text.

One row per event: the vetted original beside the plain-language draft, plus the
verified further-reading links and the drafter's own note where a judgement was
made. Editing the columns marked in blue and returning the file is enough to
update the app; apply_basic_review.py reads them straight back.
"""
import csv, json, glob, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR  = PatternFill("solid", fgColor="0B3552")
EDIT = PatternFill("solid", fgColor="E8F3FB")   # blue = you may edit this
LOCK = PatternFill("solid", fgColor="F2F2F2")   # grey = reference, do not edit
FLAG = PatternFill("solid", fgColor="FFF3D6")
THIN = Border(*(Side(style="thin", color="D0D0D0"),)*4)

rows = {r["event_id"]: r for r in
        csv.DictReader(open("coral_reef_events_master_basic.csv", encoding="utf-8-sig"))}
drafts = {}
for f in sorted(glob.glob("basic/out/batch*.json"),
                key=lambda p: int(re.search(r"\d+", p.split("/")[-1]).group())):
    for o in json.load(open(f)):
        drafts[o["event_id"]] = o

wb = Workbook()

# ------------------------------------------------------------------ read me
ws = wb.active; ws.title = "Read me"
LINES = [
 ("How to use this file", True),
 ("", False),
 ("One row per event on the timeline, on the 'Text' sheet. The white/grey columns are the", False),
 ("existing advanced text, for reference. The blue columns are the plain-language version", False),
 ("drafted for readers around 11 to 14 years old. Edit the blue columns only, save, and send", False),
 ("the file back; the app rebuilds straight from them.", False),
 ("", False),
 ("What the plain-language draft is", True),
 ("It is a restatement of your existing vetted text in simpler words. No facts were added to", False),
 ("it, no new sources were consulted, and no numbers appear in it that are not in the original.", False),
 ("", False),
 ("How it was checked", True),
 ("Three independent reviewers read all 175 rewrites against their originals looking only for", False),
 ("meaning drift: added claims, dropped qualifiers, hedges hardened into certainties, invented", False),
 ("causation, and projections written as though they had already happened.", False),
 ("They found 18 problems. None would change what a reader believes. 5 broadened or hardened a", False),
 ("claim, 13 were imprecise but not misleading. All 18 corrections are already applied here.", False),
 ("A separate mechanical check confirmed that no number appears in the plain text that is not", False),
 ("in the original: 0 of 175 events failed.", False),
 ("", False),
 ("Reading level", True),
 ("Flesch-Kincaid grade, which is roughly the US school year a reader needs. Ages 11 to 14 is", False),
 ("about grade 6 to 9.", False),
 ("    original text     median grade 13.8    14 of 175 events at or below grade 9", False),
 ("    plain-language    median grade 8.2    113 of 175 events at or below grade 9", False),
 ("", False),
 ("Further reading links", True),
 ("In basic mode the journal citations are hidden, and where a genuinely suitable page exists", False),
 ("the event links to that instead. Every link was fetched and read before it was accepted.", False),
 ("129 of 175 events have one. 46 have none, deliberately: for those events no general-audience", False),
 ("page actually covers the subject, and a weak link is worse than no link.", False),
 ("If you want to add one, put the page title and URL in the two blue link columns.", False),
 ("", False),
 ("The 'Notes' column", True),
 ("Where the drafter made a judgement call, they said so. 83 events carry a note. Those are the", False),
 ("rows worth reading first if you are short of time.", False),
]
for i, (t, bold) in enumerate(LINES, start=1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = Font(name=FONT, size=12 if bold else 10.5, bold=bold,
                  color="0B3552" if bold else "202020")
ws.column_dimensions["A"].width = 104
ws.sheet_view.showGridLines = False

# ------------------------------------------------------------------ text
ws = wb.create_sheet("Text")
COLS = [("Event", 10, "lock"), ("Date", 14, "lock"), ("Category", 22, "lock"),
        ("ORIGINAL headline", 44, "lock"), ("PLAIN headline", 44, "edit"),
        ("ORIGINAL what happened", 54, "lock"), ("PLAIN what happened", 54, "edit"),
        ("ORIGINAL why it matters", 42, "lock"), ("PLAIN why it matters", 42, "edit"),
        ("Link 1 title", 34, "edit"), ("Link 1 URL", 40, "edit"), ("Link 1 publisher", 26, "edit"),
        ("Link 2 title", 34, "edit"), ("Link 2 URL", 40, "edit"), ("Link 2 publisher", 26, "edit"),
        ("Drafter's note", 46, "lock"), ("Reading grade", 11, "lock")]
for j, (name, w, _) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR; c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "D2"

VOW = re.compile(r"[aeiouy]+")
def syl(w):
    w = w.lower().strip(".,;:!?()\"'")
    if not w: return 0
    n = len(VOW.findall(w))
    if w.endswith("e") and not w.endswith(("le","ee","ye")) and n > 1: n -= 1
    return max(1, n)
def fk(text):
    sents = [x for x in re.split(r"[.!?]+", text) if x.strip()]
    words = re.findall(r"[A-Za-z][A-Za-z'-]*", text)
    if not sents or not words: return 0.0
    return round(0.39*len(words)/len(sents) + 11.8*sum(syl(w) for w in words)/len(words) - 15.59, 1)

for i, (eid, r) in enumerate(rows.items(), start=2):
    d = drafts[eid]
    links = json.loads(r["basic_links"]) if r["basic_links"] else []
    l1 = links[0] if len(links) > 0 else {}
    l2 = links[1] if len(links) > 1 else {}
    vals = [eid, r["display_date"], r["category"],
            r["headline"], d["basic_headline"],
            r["significance"], d["basic_significance"],
            r["why_it_matters"], d["basic_why"],
            l1.get("t",""), l1.get("u",""), l1.get("p",""),
            l2.get("t",""), l2.get("u",""), l2.get("p",""),
            d.get("note",""), fk(d["basic_significance"] + " " + d["basic_why"])]
    for j, (v, (_, _, kind)) in enumerate(zip(vals, COLS), start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=j > 3)
        c.fill = EDIT if kind == "edit" else LOCK
        c.border = THIN
    if d.get("note"):
        ws.cell(row=i, column=16).fill = FLAG
    if eid.startswith("PJ"):
        ws.cell(row=i, column=1).font = Font(name=FONT, size=10, bold=True, color="1F5FA8")
    ws.row_dimensions[i].height = 62
ws.auto_filter.ref = f"A1:Q{len(rows)+1}"

# ------------------------------------------------------------------ links
ws = wb.create_sheet("Links used")
LK = json.loads(Path("basic/links.json").read_text())
for j, (name, w) in enumerate([("Topic", 40), ("Page title", 56), ("Publisher", 42),
                               ("URL", 62), ("Events using it", 46)], start=1):
    c = ws.cell(row=1, column=j, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF"); c.fill = HDR
    ws.column_dimensions[get_column_letter(j)].width = w
r = 2
for tid, t in LK["topics"].items():
    used = sorted(e for e, ts in LK["events"].items() if tid in ts)
    for l in t["links"]:
        for j, v in enumerate([t["label"], l["title"], l["publisher"], l["url"],
                               ", ".join(used) or "none"], start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=FONT, size=10); c.alignment = Alignment(vertical="top", wrap_text=j != 4)
            c.border = THIN
        r += 1
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:E{r-1}"

# ------------------------------------------------------------------ no link
ws = wb.create_sheet("No link, and why")
ws["A1"] = "Events with no further-reading link"
ws["A1"].font = Font(name=FONT, size=12, bold=True, color="0B3552")
ws["A2"] = ("These 46 events have no general-audience page that genuinely covers their subject. "
            "They show no link at all in basic mode rather than a weak one. Add a title and URL "
            "on the 'Text' sheet if you know of a good page.")
ws["A2"].font = Font(name=FONT, size=10); ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:D2"); ws.row_dimensions[2].height = 34
for j, (name, w) in enumerate([("Event", 10), ("Date", 14), ("Headline", 70), ("Category", 26)], start=1):
    c = ws.cell(row=4, column=j, value=name)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF"); c.fill = HDR
    ws.column_dimensions[get_column_letter(j)].width = w
r = 5
for eid, row in rows.items():
    if row["basic_links"]: continue
    for j, v in enumerate([eid, row["display_date"], row["headline"], row["category"]], start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(name=FONT, size=10); c.alignment = Alignment(vertical="top", wrap_text=j == 3)
        c.border = THIN
    r += 1
ws.freeze_panes = "A5"

out = "coral_timeline_basic_text_REVIEW.xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"  Text sheet          : {len(rows)} events")
print(f"  Links used sheet    : {r if False else sum(len(t['links']) for t in LK['topics'].values())} verified pages")
print(f"  No link sheet       : {sum(1 for x in rows.values() if not x['basic_links'])} events")
